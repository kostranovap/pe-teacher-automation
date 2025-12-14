from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime
from app import db
from app.models import (User, Faculty, Specialty, EducationForm, Group, Student,
                        Module, Theme, Standard, StandardScale, Attendance,
                        StandardResult, Assignment, Statement)
from app.utils import allowed_file, get_unique_filename
import os
from openpyxl import load_workbook

bp = Blueprint('admin', __name__, url_prefix='/admin')


def admin_required(f):
    """Декоратор для проверки прав администратора"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Доступ запрещен. Требуются права администратора.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function


# ===================== ПАНЕЛЬ АДМИНИСТРАТОРА =====================

@bp.route('/')
@login_required
@admin_required
def dashboard():
    """Главная панель администратора"""
    stats = {
        'users_count': User.query.count(),
        'teachers_count': User.query.filter_by(role='teacher').count(),
        'faculties_count': Faculty.query.count(),
        'specialties_count': Specialty.query.count(),
        'groups_count': Group.query.count(),
        'students_count': Student.query.count(),
        'modules_count': Module.query.count(),
        'themes_count': Theme.query.count(),
        'standards_count': Standard.query.count(),
        'active_standards_count': Standard.query.filter_by(is_active=True).count()
    }
    
    # Последние пользователи
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    
    # Последние группы
    recent_groups = Group.query.order_by(Group.created_at.desc()).limit(5).all()
    
    return render_template('admin/dashboard.html',
                         stats=stats,
                         recent_users=recent_users,
                         recent_groups=recent_groups)


# ===================== УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ =====================

@bp.route('/users')
@login_required
@admin_required
def users():
    """Список пользователей"""
    role_filter = request.args.get('role')
    is_active = request.args.get('is_active')
    
    query = User.query
    
    if role_filter:
        query = query.filter_by(role=role_filter)
    if is_active is not None:
        query = query.filter_by(is_active=is_active.lower() == 'true')
    
    users = query.order_by(User.created_at.desc()).all()
    
    return render_template('admin/users.html', users=users)


@bp.route('/users/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    """Создать пользователя"""
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        role = request.form.get('role')
        is_active = request.form.get('is_active') == 'on'
        
        # Валидация
        if not all([email, password, full_name, role]):
            flash('Все обязательные поля должны быть заполнены.', 'danger')
            return render_template('admin/user_form.html')
        
        if role not in ['admin', 'teacher', 'department_head']:
            flash('Неверная роль пользователя.', 'danger')
            return render_template('admin/user_form.html')
        
        if User.query.filter_by(email=email).first():
            flash('Пользователь с таким email уже существует.', 'danger')
            return render_template('admin/user_form.html')
        
        # Создать пользователя
        user = User(
            email=email,
            full_name=full_name,
            role=role,
            is_active=is_active
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash(f'Пользователь "{full_name}" успешно создан.', 'success')
        return redirect(url_for('admin.users'))
    
    return render_template('admin/user_form.html', user=None)


@bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    """Редактировать пользователя"""
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        email = request.form.get('email')
        full_name = request.form.get('full_name')
        role = request.form.get('role')
        is_active = request.form.get('is_active') == 'on'
        new_password = request.form.get('new_password')
        
        # Валидация
        if not all([email, full_name, role]):
            flash('Все обязательные поля должны быть заполнены.', 'danger')
            return render_template('admin/user_form.html', user=user)
        
        if role not in ['admin', 'teacher', 'department_head']:
            flash('Неверная роль пользователя.', 'danger')
            return render_template('admin/user_form.html', user=user)
        
        # Проверка email на уникальность
        existing = User.query.filter_by(email=email).first()
        if existing and existing.id != user_id:
            flash('Email уже используется другим пользователем.', 'danger')
            return render_template('admin/user_form.html', user=user)
        
        # Обновить данные
        user.email = email
        user.full_name = full_name
        user.role = role
        user.is_active = is_active
        
        if new_password:
            user.set_password(new_password)
        
        db.session.commit()
        
        flash(f'Пользователь "{full_name}" успешно обновлен.', 'success')
        return redirect(url_for('admin.users'))
    
    return render_template('admin/user_form.html', user=user)


@bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    """Удалить пользователя"""
    if user_id == current_user.id:
        flash('Нельзя удалить самого себя.', 'danger')
        return redirect(url_for('admin.users'))
    
    user = User.query.get_or_404(user_id)
    full_name = user.full_name
    
    db.session.delete(user)
    db.session.commit()
    
    flash(f'Пользователь "{full_name}" успешно удален.', 'success')
    return redirect(url_for('admin.users'))


# ===================== ФАКУЛЬТЕТЫ =====================

@bp.route('/faculties')
@login_required
@admin_required
def faculties():
    """Список факультетов"""
    faculties = Faculty.query.all()
    return render_template('admin/faculties.html', faculties=faculties)


@bp.route('/faculties/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_faculty():
    """Создать факультет"""
    if request.method == 'POST':
        code = request.form.get('code')
        name = request.form.get('name')
        
        if not all([code, name]):
            flash('Код и название обязательны.', 'danger')
            return render_template('admin/faculty_form.html')
        
        if Faculty.query.filter_by(code=code).first():
            flash('Факультет с таким кодом уже существует.', 'danger')
            return render_template('admin/faculty_form.html')
        
        faculty = Faculty(code=code, name=name)
        db.session.add(faculty)
        db.session.commit()
        
        flash(f'Факультет "{name}" успешно создан.', 'success')
        return redirect(url_for('admin.faculties'))
    
    return render_template('admin/faculty_form.html', faculty=None)


@bp.route('/faculties/<int:faculty_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_faculty(faculty_id):
    """Редактировать факультет"""
    faculty = Faculty.query.get_or_404(faculty_id)
    
    if request.method == 'POST':
        code = request.form.get('code')
        name = request.form.get('name')
        
        if not all([code, name]):
            flash('Код и название обязательны.', 'danger')
            return render_template('admin/faculty_form.html', faculty=faculty)
        
        # Проверка кода на уникальность
        existing = Faculty.query.filter_by(code=code).first()
        if existing and existing.id != faculty_id:
            flash('Код уже используется.', 'danger')
            return render_template('admin/faculty_form.html', faculty=faculty)
        
        faculty.code = code
        faculty.name = name
        db.session.commit()
        
        flash(f'Факультет "{name}" успешно обновлен.', 'success')
        return redirect(url_for('admin.faculties'))
    
    return render_template('admin/faculty_form.html', faculty=faculty)


@bp.route('/faculties/<int:faculty_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_faculty(faculty_id):
    """Удалить факультет"""
    faculty = Faculty.query.get_or_404(faculty_id)
    
    if faculty.specialties.count() > 0:
        flash('Нельзя удалить факультет с привязанными специальностями.', 'danger')
        return redirect(url_for('admin.faculties'))
    
    name = faculty.name
    db.session.delete(faculty)
    db.session.commit()
    
    flash(f'Факультет "{name}" успешно удален.', 'success')
    return redirect(url_for('admin.faculties'))


# ===================== СПЕЦИАЛЬНОСТИ =====================

@bp.route('/specialties')
@login_required
@admin_required
def specialties():
    """Список специальностей"""
    faculty_id = request.args.get('faculty_id', type=int)
    
    query = Specialty.query
    if faculty_id:
        query = query.filter_by(faculty_id=faculty_id)
    
    specialties = query.all()
    faculties = Faculty.query.all()
    
    return render_template('admin/specialties.html',
                         specialties=specialties,
                         faculties=faculties,
                         selected_faculty_id=faculty_id)


@bp.route('/specialties/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_specialty():
    """Создать специальность"""
    faculties = Faculty.query.all()
    
    if request.method == 'POST':
        code = request.form.get('code')
        name = request.form.get('name')
        faculty_id = request.form.get('faculty_id', type=int)
        
        if not all([code, name, faculty_id]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/specialty_form.html',
                                 specialty=None,
                                 faculties=faculties)
        
        if Specialty.query.filter_by(code=code).first():
            flash('Специальность с таким кодом уже существует.', 'danger')
            return render_template('admin/specialty_form.html',
                                 specialty=None,
                                 faculties=faculties)
        
        if not Faculty.query.get(faculty_id):
            flash('Факультет не найден.', 'danger')
            return render_template('admin/specialty_form.html',
                                 specialty=None,
                                 faculties=faculties)
        
        specialty = Specialty(code=code, name=name, faculty_id=faculty_id)
        db.session.add(specialty)
        db.session.commit()
        
        flash(f'Специальность "{name}" успешно создана.', 'success')
        return redirect(url_for('admin.specialties'))
    
    return render_template('admin/specialty_form.html',
                         specialty=None,
                         faculties=faculties)


@bp.route('/specialties/<int:specialty_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_specialty(specialty_id):
    """Редактировать специальность"""
    specialty = Specialty.query.get_or_404(specialty_id)
    faculties = Faculty.query.all()
    
    if request.method == 'POST':
        code = request.form.get('code')
        name = request.form.get('name')
        faculty_id = request.form.get('faculty_id', type=int)
        
        if not all([code, name, faculty_id]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/specialty_form.html',
                                 specialty=specialty,
                                 faculties=faculties)
        
        # Проверка кода на уникальность
        existing = Specialty.query.filter_by(code=code).first()
        if existing and existing.id != specialty_id:
            flash('Код уже используется.', 'danger')
            return render_template('admin/specialty_form.html',
                                 specialty=specialty,
                                 faculties=faculties)
        
        if not Faculty.query.get(faculty_id):
            flash('Факультет не найден.', 'danger')
            return render_template('admin/specialty_form.html',
                                 specialty=specialty,
                                 faculties=faculties)
        
        specialty.code = code
        specialty.name = name
        specialty.faculty_id = faculty_id
        db.session.commit()
        
        flash(f'Специальность "{name}" успешно обновлена.', 'success')
        return redirect(url_for('admin.specialties'))
    
    return render_template('admin/specialty_form.html',
                         specialty=specialty,
                         faculties=faculties)


@bp.route('/specialties/<int:specialty_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_specialty(specialty_id):
    """Удалить специальность"""
    specialty = Specialty.query.get_or_404(specialty_id)
    
    if specialty.groups.count() > 0:
        flash('Нельзя удалить специальность с привязанными группами.', 'danger')
        return redirect(url_for('admin.specialties'))
    
    name = specialty.name
    db.session.delete(specialty)
    db.session.commit()
    
    flash(f'Специальность "{name}" успешно удалена.', 'success')
    return redirect(url_for('admin.specialties'))


# ===================== ФОРМЫ ОБУЧЕНИЯ =====================

@bp.route('/education-forms')
@login_required
@admin_required
def education_forms():
    """Список форм обучения"""
    forms = EducationForm.query.all()
    return render_template('admin/education_forms.html', forms=forms)


@bp.route('/education-forms/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_education_form():
    """Создать форму обучения"""
    if request.method == 'POST':
        name = request.form.get('name')
        duration_years = request.form.get('duration_years', type=float)
        
        if not all([name, duration_years]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/education_form_form.html')
        
        if EducationForm.query.filter_by(name=name).first():
            flash('Форма обучения с таким названием уже существует.', 'danger')
            return render_template('admin/education_form_form.html')
        
        form = EducationForm(name=name, duration_years=duration_years)
        db.session.add(form)
        db.session.commit()
        
        flash(f'Форма обучения "{name}" успешно создана.', 'success')
        return redirect(url_for('admin.education_forms'))
    
    return render_template('admin/education_form_form.html', form=None)


@bp.route('/education-forms/<int:form_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_education_form(form_id):
    """Удалить форму обучения"""
    form = EducationForm.query.get_or_404(form_id)
    
    if form.groups.count() > 0:
        flash('Нельзя удалить форму обучения с привязанными группами.', 'danger')
        return redirect(url_for('admin.education_forms'))
    
    name = form.name
    db.session.delete(form)
    db.session.commit()
    
    flash(f'Форма обучения "{name}" успешно удалена.', 'success')
    return redirect(url_for('admin.education_forms'))


# ===================== ГРУППЫ (С DEBUG) =====================

@bp.route('/groups')
@login_required
@admin_required
def groups():
    """Список групп"""
    print("\n" + "="*50)
    print("DEBUG: groups() function called!")
    print("Rendering: admin/groups.html")
    print("="*50 + "\n")
    
    teacher_id = request.args.get('teacher_id', type=int)
    course = request.args.get('course', type=int)
    
    query = Group.query
    if teacher_id:
        query = query.filter_by(teacher_id=teacher_id)
    if course:
        query = query.filter_by(course=course)
    
    groups = query.order_by(Group.name).all()
    teachers = User.query.filter_by(role='teacher').all()
    
    return render_template('admin/groups.html',
                         groups=groups,
                         teachers=teachers,
                         selected_teacher_id=teacher_id,
                         selected_course=course)


@bp.route('/groups/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_group():
    """Создать группу"""
    print("\n" + "="*50)
    print("DEBUG: create_group() function called!")
    print("Rendering: admin/group_form.html")
    print("="*50 + "\n")
    
    specialties = Specialty.query.all()
    education_forms = EducationForm.query.all()
    teachers = User.query.filter_by(role='teacher').all()
    
    if request.method == 'POST':
        name = request.form.get('name')
        course = request.form.get('course', type=int)
        semester = request.form.get('semester', type=int)
        specialty_id = request.form.get('specialty_id', type=int)
        education_form_id = request.form.get('education_form_id', type=int)
        teacher_id = request.form.get('teacher_id', type=int) or None
        
        if not all([name, course, semester, specialty_id, education_form_id]):
            flash('Все обязательные поля должны быть заполнены.', 'danger')
            return render_template('admin/group_form.html',
                                 group=None,
                                 specialties=specialties,
                                 education_forms=education_forms,
                                 teachers=teachers)
        
        if Group.query.filter_by(name=name).first():
            flash('Группа с таким названием уже существует.', 'danger')
            return render_template('admin/group_form.html',
                                 group=None,
                                 specialties=specialties,
                                 education_forms=education_forms,
                                 teachers=teachers)
        
        group = Group(
            name=name,
            course=course,
            semester=semester,
            specialty_id=specialty_id,
            education_form_id=education_form_id,
            teacher_id=teacher_id
        )
        db.session.add(group)
        db.session.commit()
        
        flash(f'Группа "{name}" успешно создана.', 'success')
        return redirect(url_for('admin.groups'))
    
    return render_template('admin/group_form.html',
                         group=None,
                         specialties=specialties,
                         education_forms=education_forms,
                         teachers=teachers)


@bp.route('/groups/<int:group_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_group(group_id):
    """Редактировать группу"""
    group = Group.query.get_or_404(group_id)
    specialties = Specialty.query.all()
    education_forms = EducationForm.query.all()
    teachers = User.query.filter_by(role='teacher').all()
    
    if request.method == 'POST':
        name = request.form.get('name')
        course = request.form.get('course', type=int)
        semester = request.form.get('semester', type=int)
        specialty_id = request.form.get('specialty_id', type=int)
        education_form_id = request.form.get('education_form_id', type=int)
        teacher_id = request.form.get('teacher_id', type=int) or None
        
        if not all([name, course, semester, specialty_id, education_form_id]):
            flash('Все обязательные поля должны быть заполнены.', 'danger')
            return render_template('admin/group_form.html',
                                 group=group,
                                 specialties=specialties,
                                 education_forms=education_forms,
                                 teachers=teachers)
        
        # Проверка названия на уникальность
        existing = Group.query.filter_by(name=name).first()
        if existing and existing.id != group_id:
            flash('Название уже используется.', 'danger')
            return render_template('admin/group_form.html',
                                 group=group,
                                 specialties=specialties,
                                 education_forms=education_forms,
                                 teachers=teachers)
        
        group.name = name
        group.course = course
        group.semester = semester
        group.specialty_id = specialty_id
        group.education_form_id = education_form_id
        group.teacher_id = teacher_id
        db.session.commit()
        
        flash(f'Группа "{name}" успешно обновлена.', 'success')
        return redirect(url_for('admin.groups'))
    
    return render_template('admin/group_form.html',
                         group=group,
                         specialties=specialties,
                         education_forms=education_forms,
                         teachers=teachers)


@bp.route('/groups/<int:group_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_group(group_id):
    """Удалить группу"""
    group = Group.query.get_or_404(group_id)
    
    if group.students.count() > 0:
        flash('Нельзя удалить группу со студентами.', 'danger')
        return redirect(url_for('admin.groups'))
    
    name = group.name
    db.session.delete(group)
    db.session.commit()
    
    flash(f'Группа "{name}" успешно удалена.', 'success')
    return redirect(url_for('admin.groups'))

# ===================== ИМПОРТ СТУДЕНТОВ =====================

@bp.route('/students/import', methods=['GET', 'POST'])
@login_required
@admin_required
def import_students():
    """Импорт студентов из Excel"""
    groups = Group.query.order_by(Group.name).all()
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не найден.', 'danger')
            return render_template('admin/import_students.html', groups=groups)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('Файл не выбран.', 'danger')
            return render_template('admin/import_students.html', groups=groups)
        
        if not allowed_file(file.filename, {'xlsx', 'xls'}):
            flash('Неверный формат файла. Используйте .xlsx или .xls', 'danger')
            return render_template('admin/import_students.html', groups=groups)
        
        try:
            # Сохранить файл временно
            filename = get_unique_filename(file.filename)
            filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Загрузить Excel
            workbook = load_workbook(filepath)
            sheet = workbook.active
            
            imported = 0
            errors = []
            
            # Пропустить заголовок
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Пустая строка
                    continue
                
                try:
                    full_name = row[0]
                    student_number = row[1]
                    group_name = row[2]
                    gender = row[3]
                    birth_date = row[4] if len(row) > 4 else None
                    medical_group = row[5] if len(row) > 5 else 'основная'
                    
                    # Проверка обязательных полей
                    if not all([full_name, student_number, group_name, gender]):
                        errors.append(f'Строка {row_num}: отсутствуют обязательные поля')
                        continue
                    
                    # Проверка пола
                    if gender.lower() not in ['м', 'ж', 'male', 'female', 'мужской', 'женский']:
                        errors.append(f'Строка {row_num}: неверный пол')
                        continue
                    
                    # Нормализация пола
                    if gender.lower() in ['м', 'male', 'мужской']:
                        gender = 'male'
                    else:
                        gender = 'female'
                    
                    # Найти группу
                    group = Group.query.filter_by(name=group_name).first()
                    if not group:
                        errors.append(f'Строка {row_num}: группа "{group_name}" не найдена')
                        continue
                    
                    # Проверка на дубликат
                    if Student.query.filter_by(student_number=student_number).first():
                        errors.append(f'Строка {row_num}: студент с номером {student_number} уже существует')
                        continue
                    
                    # Создать студента
                    student = Student(
                        full_name=full_name,
                        student_number=student_number,
                        gender=gender,
                        birth_date=birth_date if isinstance(birth_date, datetime) else None,
                        medical_group=medical_group,
                        group_id=group.id
                    )
                    db.session.add(student)
                    imported += 1
                    
                except Exception as e:
                    errors.append(f'Строка {row_num}: {str(e)}')
            
            db.session.commit()
            
            # Удалить временный файл
            os.remove(filepath)
            
            if imported > 0:
                flash(f'Импорт завершен. Добавлено студентов: {imported}', 'success')
            
            if errors:
                flash(f'Обнаружено ошибок: {len(errors)}', 'warning')
                for error in errors[:10]:  # Показать первые 10 ошибок
                    flash(error, 'warning')
            
            return redirect(url_for('admin.import_students'))
            
        except Exception as e:
            flash(f'Ошибка импорта: {str(e)}', 'danger')
            return render_template('admin/import_students.html', groups=groups)
    
    return render_template('admin/import_students.html', groups=groups)


# ===================== МОДУЛИ И ТЕМЫ =====================

@bp.route('/modules')
@login_required
@admin_required
def modules():
    """Список модулей и тем"""
    modules = Module.query.order_by(Module.number).all()
    return render_template('admin/modules.html', modules=modules)


@bp.route('/modules/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_module():
    """Создать модуль"""
    if request.method == 'POST':
        number = request.form.get('number', type=int)
        name = request.form.get('name')
        max_points = request.form.get('max_points', type=int)
        
        if not all([number, name, max_points]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/module_form.html', module=None)
        
        if Module.query.filter_by(number=number).first():
            flash('Модуль с таким номером уже существует.', 'danger')
            return render_template('admin/module_form.html', module=None)
        
        module = Module(number=number, name=name, max_points=max_points)
        db.session.add(module)
        db.session.commit()
        
        flash(f'Модуль "{name}" успешно создан.', 'success')
        return redirect(url_for('admin.modules'))
    
    return render_template('admin/module_form.html', module=None)


@bp.route('/themes/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_theme():
    """Создать тему"""
    modules = Module.query.order_by(Module.number).all()
    
    if request.method == 'POST':
        name = request.form.get('name')
        module_id = request.form.get('module_id', type=int)
        max_points = request.form.get('max_points', type=int)
        
        if not all([name, module_id, max_points]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/theme_form.html',
                                 theme=None,
                                 modules=modules)
        
        if not Module.query.get(module_id):
            flash('Модуль не найден.', 'danger')
            return render_template('admin/theme_form.html',
                                 theme=None,
                                 modules=modules)
        
        theme = Theme(name=name, module_id=module_id, max_points=max_points)
        db.session.add(theme)
        db.session.commit()
        
        flash(f'Тема "{name}" успешно создана.', 'success')
        return redirect(url_for('admin.modules'))
    
    return render_template('admin/theme_form.html', theme=None, modules=modules)

@bp.route('/modules/<int:module_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_module(module_id):
    """Редактировать модуль"""
    module = Module.query.get_or_404(module_id)
    
    if request.method == 'POST':
        number = request.form.get('number', type=int)
        name = request.form.get('name')
        max_points = request.form.get('max_points', type=int)
        
        if not all([number, name, max_points]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/module_form.html', module=module)
        
        # Проверка номера на уникальность
        existing = Module.query.filter_by(number=number).first()
        if existing and existing.id != module_id:
            flash('Модуль с таким номером уже существует.', 'danger')
            return render_template('admin/module_form.html', module=module)
        
        module.number = number
        module.name = name
        module.max_points = max_points
        db.session.commit()
        
        flash(f'Модуль "{name}" успешно обновлен.', 'success')
        return redirect(url_for('admin.modules'))
    
    return render_template('admin/module_form.html', module=module)


@bp.route('/modules/<int:module_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_module(module_id):
    """Удалить модуль (каскадное удаление всех тем)"""
    module = Module.query.get_or_404(module_id)
    name = module.name
    
    # Подсчет удаляемых объектов
    themes_count = module.themes.count()
    standards_count = 0
    
    # Каскадное удаление: сначала удалить все темы и их нормативы
    for theme in module.themes.all():
        standards_count += theme.standards.count()
        # Удалить все нормативы темы
        for standard in theme.standards.all():
            # Удалить оценочные шкалы норматива
            StandardScale.query.filter_by(standard_id=standard.id).delete()
            # Удалить результаты норматива
            StandardResult.query.filter_by(standard_id=standard.id).delete()
            db.session.delete(standard)
        # Удалить тему
        db.session.delete(theme)
    
    # Удалить модуль
    db.session.delete(module)
    db.session.commit()
    
    # Информативное сообщение
    message = f'Модуль "{name}" успешно удален.'
    if themes_count > 0:
        message += f' Удалено тем: {themes_count}.'
    if standards_count > 0:
        message += f' Удалено нормативов: {standards_count}.'
    
    flash(message, 'success')
    return redirect(url_for('admin.modules'))


@bp.route('/themes/<int:theme_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_theme(theme_id):
    """Редактировать тему"""
    theme = Theme.query.get_or_404(theme_id)
    modules = Module.query.order_by(Module.number).all()
    
    if request.method == 'POST':
        name = request.form.get('name')
        module_id = request.form.get('module_id', type=int)
        max_points = request.form.get('max_points', type=int)
        
        if not all([name, module_id, max_points]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/theme_form.html',
                                 theme=theme,
                                 modules=modules)
        
        if not Module.query.get(module_id):
            flash('Модуль не найден.', 'danger')
            return render_template('admin/theme_form.html',
                                 theme=theme,
                                 modules=modules)
        
        theme.name = name
        theme.module_id = module_id
        theme.max_points = max_points
        db.session.commit()
        
        flash(f'Тема "{name}" успешно обновлена.', 'success')
        return redirect(url_for('admin.modules'))
    
    return render_template('admin/theme_form.html', theme=theme, modules=modules)


@bp.route('/themes/<int:theme_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_theme(theme_id):
    """Удалить тему (каскадное удаление всех нормативов)"""
    theme = Theme.query.get_or_404(theme_id)
    name = theme.name
    
    # Подсчет удаляемых объектов
    standards_count = theme.standards.count()
    
    # Каскадное удаление: удалить все нормативы темы
    for standard in theme.standards.all():
        # Удалить оценочные шкалы норматива
        StandardScale.query.filter_by(standard_id=standard.id).delete()
        # Удалить результаты норматива
        StandardResult.query.filter_by(standard_id=standard.id).delete()
        db.session.delete(standard)
    
    # Удалить тему
    db.session.delete(theme)
    db.session.commit()
    
    # Информативное сообщение
    message = f'Тема "{name}" успешно удалена.'
    if standards_count > 0:
        message += f' Удалено нормативов: {standards_count}.'
    
    flash(message, 'success')
    return redirect(url_for('admin.modules'))

# ===================== НОРМАТИВЫ =====================

@bp.route('/standards')
@login_required
@admin_required
def standards():
    """Список нормативов"""
    theme_id = request.args.get('theme_id', type=int)
    is_active = request.args.get('is_active')
    
    query = Standard.query
    if theme_id:
        query = query.filter_by(theme_id=theme_id)
    if is_active is not None:
        query = query.filter_by(is_active=is_active.lower() == 'true')
    
    standards = query.all()
    themes = Theme.query.all()
    
    return render_template('admin/standards.html',
                         standards=standards,
                         themes=themes,
                         selected_theme_id=theme_id)


@bp.route('/standards/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_standard():
    """Создать норматив"""
    themes = Theme.query.all()
    
    if request.method == 'POST':
        name = request.form.get('name')
        theme_id = request.form.get('theme_id', type=int)
        unit = request.form.get('unit')
        comparison_type = request.form.get('comparison_type')
        is_active = request.form.get('is_active') == 'on'
        
        if not all([name, theme_id, unit, comparison_type]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/standard_form.html',
                                 standard=None,
                                 themes=themes)
        
        if comparison_type not in ['less_better', 'more_better']:
            flash('Неверный тип сравнения.', 'danger')
            return render_template('admin/standard_form.html',
                                 standard=None,
                                 themes=themes)
        
        if not Theme.query.get(theme_id):
            flash('Тема не найдена.', 'danger')
            return render_template('admin/standard_form.html',
                                 standard=None,
                                 themes=themes)
        
        standard = Standard(
            name=name,
            theme_id=theme_id,
            unit=unit,
            comparison_type=comparison_type,
            is_active=is_active
        )
        db.session.add(standard)
        db.session.commit()
        
        flash(f'Норматив "{name}" успешно создан.', 'success')
        return redirect(url_for('admin.standards'))
    
    return render_template('admin/standard_form.html',
                         standard=None,
                         themes=themes)


@bp.route('/standards/<int:standard_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_standard(standard_id):
    """Редактировать норматив"""
    standard = Standard.query.get_or_404(standard_id)
    themes = Theme.query.all()
    
    if request.method == 'POST':
        name = request.form.get('name')
        unit = request.form.get('unit')
        comparison_type = request.form.get('comparison_type')
        is_active = request.form.get('is_active') == 'on'
        
        if not all([name, unit, comparison_type]):
            flash('Все поля обязательны.', 'danger')
            return render_template('admin/standard_form.html',
                                 standard=standard,
                                 themes=themes)
        
        if comparison_type not in ['less_better', 'more_better']:
            flash('Неверный тип сравнения.', 'danger')
            return render_template('admin/standard_form.html',
                                 standard=standard,
                                 themes=themes)
        
        standard.name = name
        standard.unit = unit
        standard.comparison_type = comparison_type
        standard.is_active = is_active
        db.session.commit()
        
        flash(f'Норматив "{name}" успешно обновлен.', 'success')
        return redirect(url_for('admin.standards'))
    
    return render_template('admin/standard_form.html',
                         standard=standard,
                         themes=themes)


@bp.route('/standards/<int:standard_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_standard(standard_id):
    """Удалить норматив"""
    standard = Standard.query.get_or_404(standard_id)
    
    if standard.results.count() > 0:
        flash('Нельзя удалить норматив с привязанными результатами.', 'danger')
        return redirect(url_for('admin.standards'))
    
    name = standard.name
    db.session.delete(standard)
    db.session.commit()
    
    flash(f'Норматив "{name}" успешно удален.', 'success')
    return redirect(url_for('admin.standards'))


# ===================== ОЦЕНОЧНЫЕ ШКАЛЫ =====================

@bp.route('/standard-scales')
@login_required
@admin_required
def standard_scales():
    """Список оценочных шкал"""
    standard_id = request.args.get('standard_id', type=int)
    
    query = StandardScale.query
    if standard_id:
        query = query.filter_by(standard_id=standard_id)
    
    scales = query.order_by(StandardScale.standard_id, 
                           StandardScale.gender,
                           StandardScale.points.desc()).all()
    standards = Standard.query.filter_by(is_active=True).all()
    
    return render_template('admin/standard_scales.html',
                         scales=scales,
                         standards=standards,
                         selected_standard_id=standard_id)


@bp.route('/standard-scales/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_standard_scale():
    """Создать оценочную шкалу"""
    standards = Standard.query.filter_by(is_active=True).all()
    
    if request.method == 'POST':
        standard_id = request.form.get('standard_id', type=int)
        gender = request.form.get('gender')
        points = request.form.get('points', type=int)
        min_value = request.form.get('min_value', type=float) or None
        max_value = request.form.get('max_value', type=float) or None
        
        if not all([standard_id, gender, points]):
            flash('Обязательные поля: норматив, пол, баллы.', 'danger')
            return render_template('admin/standard_scale_form.html',
                                 scale=None,
                                 standards=standards)
        
        if gender not in ['male', 'female']:
            flash('Неверный пол.', 'danger')
            return render_template('admin/standard_scale_form.html',
                                 scale=None,
                                 standards=standards)
        
        if points not in [1, 2, 3, 4, 5]:
            flash('Баллы должны быть от 1 до 5.', 'danger')
            return render_template('admin/standard_scale_form.html',
                                 scale=None,
                                 standards=standards)
        
        if not Standard.query.get(standard_id):
            flash('Норматив не найден.', 'danger')
            return render_template('admin/standard_scale_form.html',
                                 scale=None,
                                 standards=standards)
        
        scale = StandardScale(
            standard_id=standard_id,
            gender=gender,
            points=points,
            min_value=min_value,
            max_value=max_value
        )
        db.session.add(scale)
        db.session.commit()
        
        flash('Оценочная шкала успешно создана.', 'success')
        return redirect(url_for('admin.standard_scales'))
    
    return render_template('admin/standard_scale_form.html',
                         scale=None,
                         standards=standards)


@bp.route('/standard-scales/<int:scale_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_standard_scale(scale_id):
    """Редактировать оценочную шкалу"""
    scale = StandardScale.query.get_or_404(scale_id)
    standards = Standard.query.filter_by(is_active=True).all()
    
    if request.method == 'POST':
        points = request.form.get('points', type=int)
        min_value = request.form.get('min_value', type=float) or None
        max_value = request.form.get('max_value', type=float) or None
        
        if not points:
            flash('Баллы обязательны.', 'danger')
            return render_template('admin/standard_scale_form.html',
                                 scale=scale,
                                 standards=standards)
        
        if points not in [1, 2, 3, 4, 5]:
            flash('Баллы должны быть от 1 до 5.', 'danger')
            return render_template('admin/standard_scale_form.html',
                                 scale=scale,
                                 standards=standards)
        
        scale.points = points
        scale.min_value = min_value
        scale.max_value = max_value
        db.session.commit()
        
        flash('Оценочная шкала успешно обновлена.', 'success')
        return redirect(url_for('admin.standard_scales'))
    
    return render_template('admin/standard_scale_form.html',
                         scale=scale,
                         standards=standards)


@bp.route('/standard-scales/<int:scale_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_standard_scale(scale_id):
    """Удалить оценочную шкалу"""
    scale = StandardScale.query.get_or_404(scale_id)
    db.session.delete(scale)
    db.session.commit()
    
    flash('Оценочная шкала успешно удалена.', 'success')
    return redirect(url_for('admin.standard_scales'))